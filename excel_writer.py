"""Excel writing utilities.

Responsible for creating/updating the Operations sheet without overwriting data.
"""

from __future__ import annotations

from pathlib import Path
import logging

import pandas as pd
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

TARGET_SHEET = "Operations"
TARGET_COLUMNS = [
    "Date",
    "Tiers",
    "Libellé brut",
    "Montant",
    "Source PDF",
    "Date import",
    "ID opération",
]


def get_last_non_empty_row(ws) -> int:
    """Return last row index containing at least one non-empty cell.

    Why this helper exists:
    - `ws.max_row` can stay high after users clear cells manually in Excel.
    - We therefore scan rows from bottom to top and stop at the first row
      containing at least one non-empty value.

    Returns at least 1 to preserve header row semantics.
    """
    # Start from openpyxl worksheet max_row and search upwards.
    for row_idx in range(ws.max_row, 0, -1):
        row_values = ws.iter_rows(
            min_row=row_idx,
            max_row=row_idx,
            min_col=1,
            max_col=max(len(TARGET_COLUMNS), ws.max_column),
            values_only=True,
        )
        values = next(row_values)

        if any(value not in (None, "") for value in values):
            return max(1, row_idx)

    return 1


def append_operations_to_excel(excel_path: Path, new_data: pd.DataFrame) -> int:
    """Append non-duplicate operations to Excel.

    Returns:
        Number of rows appended.
    """
    if new_data.empty:
        return 0

    logger.info("Loading Excel file: %s", excel_path)
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()

    if TARGET_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TARGET_SHEET)
        for idx, col_name in enumerate(TARGET_COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=col_name)
    else:
        ws = wb[TARGET_SHEET]

    # Build existing ID set (column 7 expected by target layout).
    existing_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 7 and row[6]:
            existing_ids.add(str(row[6]))

    # Compute next insertion row based on real non-empty rows.
    next_row = get_last_non_empty_row(ws) + 1

    appended = 0
    for _, operation in new_data.iterrows():
        operation_id = str(operation["ID opération"])
        if operation_id in existing_ids:
            continue

        # Write row cell-by-cell to avoid ws.append() behavior with stale max_row.
        for col_idx, column_name in enumerate(TARGET_COLUMNS, start=1):
            ws.cell(row=next_row, column=col_idx, value=operation[column_name])

        existing_ids.add(operation_id)
        appended += 1
        next_row += 1

    # Remove default 'Sheet' only if empty and not used.
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        if wb["Sheet"]["A1"].value is None:
            del wb["Sheet"]

    wb.save(excel_path)
    logger.info("Excel updated. Appended rows: %s", appended)
    return appended